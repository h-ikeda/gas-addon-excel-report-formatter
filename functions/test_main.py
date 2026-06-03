"""Tests for the ``generate_excel`` Cloud Function.

These exercise the function through the Functions Framework's Flask test
client, which provides the request/response and application context that
``request.get_json`` and ``flask.jsonify`` rely on. The source under
``main.py`` is treated as a black box and is not modified.

The Excel template is no longer bundled with the function: callers (the GAS
frontend) read it from Google Drive and pass it as a Base64 ``template``
field in the request body. The helpers below build such a template in memory.

Where data lands in the sheet is defined in ``mapping.json``; these tests
load the same mapping so they stay in sync if the layout is adjusted there.
"""

import base64
import io
import json
import os

import pytest
from functions_framework import create_app
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection


SOURCE = os.path.join(os.path.dirname(__file__), "main.py")
MAPPING = json.load(open(os.path.join(os.path.dirname(__file__), "mapping.json"), encoding="utf-8"))
SHEET = MAPPING["sheet_name"]
BLOCK = MAPPING["room_block"]
STARTS = MAPPING["block_start_rows"]


@pytest.fixture
def client():
    app = create_app(target="generate_excel", source=SOURCE)
    return app.test_client()


def _make_template_b64(sheet_name=SHEET):
    """Build a minimal template workbook and return it Base64-encoded."""
    wb = Workbook()
    wb.active.title = sheet_name
    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def _payload(**fields):
    """Build a request body, defaulting to a valid Base64 template."""
    body = {"template": _make_template_b64()}
    body.update(fields)
    return body


def _decode_workbook(file_data):
    """Decode the base64 payload returned by the function into a workbook."""
    raw = base64.b64decode(file_data)
    return load_workbook(io.BytesIO(raw))


def _measurement(measurement_key):
    return next(m for m in BLOCK["measurements"] if m["key"] == measurement_key)


def _value_cell(room_index, measurement_key, field):
    """Resolve the cell a numeric value (diff/distance/digital_level) lands in,
    straight from the mapping, so assertions follow the config."""
    start = STARTS[room_index]
    offset = _measurement(measurement_key)["row_offset"]
    return f"{BLOCK['value_fields'][field]}{start + offset}"


def _select_cell(room_index, measurement_key):
    """Resolve the cell the select value (傾斜方向 / 測定した壁・柱) lands in."""
    start = STARTS[room_index]
    m = _measurement(measurement_key)
    return f"{m['select']['col']}{start + m['row_offset']}"


# --- CORS preflight ---------------------------------------------------------

def test_options_preflight_returns_cors_headers(client):
    resp = client.options("/")

    assert resp.status_code == 204
    assert resp.data == b""
    assert resp.headers["Access-Control-Allow-Origin"] == "*"
    assert resp.headers["Access-Control-Allow-Methods"] == "POST"
    assert resp.headers["Access-Control-Allow-Headers"] == "Content-Type"
    assert resp.headers["Access-Control-Max-Age"] == "3600"


# --- Successful generation --------------------------------------------------

def test_post_valid_data_returns_success_payload(client):
    resp = client.post("/", json=_payload(
        property_name="サンプル物件",
        rooms=[{"floor": "2", "room_name": "LDK", "measurements": {}}],
    ))

    assert resp.status_code == 200
    assert resp.headers["Access-Control-Allow-Origin"] == "*"

    body = resp.get_json()
    assert body["status"] == "success"
    assert body["fileName"] == "傾斜測定報告書.xlsx"
    assert body["fileData"]


def test_post_valid_data_writes_values_into_template(client):
    resp = client.post("/", json=_payload(
        property_name="サンプル物件",
        rooms=[{
            "floor": "2",
            "room_name": "LDK",
            "measurements": {
                # 床: 傾斜方向は S 列へ
                "floor_x": {"select": "←", "diff": "0", "distance": "2000"},
                "floor_y": {"select": "↑", "diff": "3", "distance": "1500"},
                # 壁: 測定した壁は P 列へ
                "wall_ud": {"select": "上壁", "diff": "2", "distance": "1800"},
                # 柱: 計測できなかった場合は ―
                "pillar_lr": {"select": "―"},
            },
        }],
    ))
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]

    # Common field
    assert ws[MAPPING["property_name_cell"]].value == "サンプル物件"
    # Room 0 header
    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value == 2
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value == "LDK"
    # 床 floor_x: select lands in S, numeric strings are coerced so AJ computes
    assert ws[_select_cell(0, "floor_x")].value == "←"
    assert ws[_value_cell(0, "floor_x", "diff")].value == 0
    assert ws[_value_cell(0, "floor_x", "distance")].value == 2000
    assert ws[_select_cell(0, "floor_y")].value == "↑"
    # 壁: select lands in P (not S)
    assert ws[_select_cell(0, "wall_ud")].value == "上壁"
    assert ws[_value_cell(0, "wall_ud", "diff")].value == 2
    # 柱: the ― ("couldn't measure") option is written verbatim to P
    assert ws[_select_cell(0, "pillar_lr")].value == "―"


def test_fullwidth_numbers_are_normalized_to_numeric(client):
    # Mobile IME often produces full-width digits/signs; they must be coerced to
    # real numbers so the AJ formula (=1000*AC/AG) can compute.
    resp = client.post("/", json=_payload(rooms=[{
        "floor": "２", "room_name": "LDK",
        "measurements": {"floor_x": {"diff": "－３", "distance": "１５００"}},
    }]))
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]

    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value == 2
    assert ws[_value_cell(0, "floor_x", "diff")].value == -3
    assert ws[_value_cell(0, "floor_x", "distance")].value == 1500


def test_non_dict_room_entries_are_skipped(client):
    # A malformed rooms array (containing a null) must not crash the function.
    resp = client.post("/", json=_payload(rooms=[
        None,
        {"floor": "2", "room_name": "寝室", "measurements": {}},
    ]))
    assert resp.status_code == 200
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]
    assert ws[f"{BLOCK['room_name_col']}{STARTS[1]}"].value == "寝室"


def test_rooms_not_a_list_returns_400(client):
    # A non-list ``rooms`` would crash len(); it must be rejected with a 400.
    resp = client.post("/", json=_payload(rooms=5))

    assert resp.status_code == 400
    assert resp.get_json() == {"error": "rooms must be a list"}


def test_non_dict_measurements_are_ignored(client):
    # ``measurements`` of the wrong type must not crash (regression guard).
    resp = client.post("/", json=_payload(rooms=[
        {"floor": "1", "room_name": "和室", "measurements": ["bogus"]},
    ]))
    assert resp.status_code == 200
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value == "和室"


def test_multiple_rooms_map_to_successive_blocks(client):
    rooms = [
        {"floor": "1", "room_name": "玄関", "measurements": {}},
        {"floor": "2", "room_name": "寝室", "measurements": {}},
    ]
    resp = client.post("/", json=_payload(rooms=rooms))
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]

    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value == "玄関"
    assert ws[f"{BLOCK['room_name_col']}{STARTS[1]}"].value == "寝室"


def test_too_many_rooms_returns_400(client):
    rooms = [{"floor": "1", "room_name": f"部屋{i}", "measurements": {}}
             for i in range(len(STARTS) + 1)]
    resp = client.post("/", json=_payload(rooms=rooms))

    assert resp.status_code == 400
    assert "上限" in resp.get_json()["error"]


def test_returned_file_is_a_valid_xlsx(client):
    resp = client.post("/", json=_payload(
        rooms=[{"floor": "1", "room_name": "和室", "measurements": {}}]))
    raw = base64.b64decode(resp.get_json()["fileData"])

    # XLSX files are zip archives and start with the PK signature.
    assert raw[:2] == b"PK"
    wb = _decode_workbook(resp.get_json()["fileData"])
    assert wb.sheetnames == [SHEET]


def test_empty_or_missing_fields_default_to_empty(client):
    # No rooms and no property name: nothing is written, file still produced.
    resp = client.post("/", json=_payload())
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]

    assert ws[MAPPING["property_name_cell"]].value is None
    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value is None


def test_blank_measurement_fields_leave_cells_empty(client):
    # A room with empty measurement values writes no data to those cells.
    resp = client.post("/", json=_payload(rooms=[{
        "floor": "3", "room_name": "",
        "measurements": {"floor_x": {"select": "", "diff": "", "distance": ""}},
    }]))
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]

    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value == 3
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value is None
    assert ws[_select_cell(0, "floor_x")].value is None
    assert ws[_value_cell(0, "floor_x", "diff")].value is None


# --- Sheet protection -------------------------------------------------------

def _make_protected_template_b64():
    """シート保護が有効で、最初のブロック以外のセルをアンロックしたテンプレートを作成する。

    最初のブロック（入力例）はロックされたまま、2ブロック目以降のセルと
    property_name_cell を locked=False にして記入可能とする。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.protection.sheet = True

    ws[MAPPING["property_name_cell"]].protection = Protection(locked=False)

    for start_row in STARTS[1:]:
        ws[f"{BLOCK['floor_col']}{start_row}"].protection = Protection(locked=False)
        ws[f"{BLOCK['room_name_col']}{start_row}"].protection = Protection(locked=False)
        for m in BLOCK["measurements"]:
            row = start_row + m["row_offset"]
            if "select" in m:
                ws[f"{m['select']['col']}{row}"].protection = Protection(locked=False)
            for col in BLOCK["value_fields"].values():
                ws[f"{col}{row}"].protection = Protection(locked=False)

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def test_locked_cells_in_protected_sheet_are_not_written(client):
    # シート保護が有効でロックされたセルへは書き込まない（入力例ブロックを保護）。
    resp = client.post("/", json={
        "template": _make_protected_template_b64(),
        "rooms": [{"floor": "2", "room_name": "LDK", "measurements": {}}],
    })
    assert resp.status_code == 200
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]
    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value is None
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value is None


def test_unlocked_cells_in_protected_sheet_are_written(client):
    # シート保護が有効でも locked=False のセルには通常通り書き込む。
    resp = client.post("/", json={
        "template": _make_protected_template_b64(),
        "property_name": "テスト物件",
        "rooms": [
            None,  # 最初のブロックはロック済みのためスキップ
            {"floor": "3", "room_name": "寝室", "measurements": {
                "floor_x": {"select": "←", "diff": "2", "distance": "1500"},
            }},
        ],
    })
    assert resp.status_code == 200
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]
    assert ws[MAPPING["property_name_cell"]].value == "テスト物件"
    assert ws[f"{BLOCK['floor_col']}{STARTS[1]}"].value == 3
    assert ws[f"{BLOCK['room_name_col']}{STARTS[1]}"].value == "寝室"
    assert ws[_select_cell(1, "floor_x")].value == "←"
    assert ws[_value_cell(1, "floor_x", "diff")].value == 2


def test_first_room_lands_in_a_writable_block(client):
    # 回帰テスト: 実雛形（IP_230901_11.xlsx）では入力例ブロックは block_start_rows
    # の外にあり、全ブロックが書き込み可能。block_start_rows[0] がロックされた
    # 入力例を指していると 1 部屋目が黙って失われる（保護セルへの書き込みは
    # スキップされるため）。block_start_rows の全ブロックが書き込めることを保証する。
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.protection.sheet = True
    # 物件名と全データブロックをアンロック（実雛形の状態を再現）。
    ws[MAPPING["property_name_cell"]].protection = Protection(locked=False)
    for start_row in STARTS:
        ws[f"{BLOCK['floor_col']}{start_row}"].protection = Protection(locked=False)
        ws[f"{BLOCK['room_name_col']}{start_row}"].protection = Protection(locked=False)
        for m in BLOCK["measurements"]:
            row = start_row + m["row_offset"]
            if "select" in m:
                ws[f"{m['select']['col']}{row}"].protection = Protection(locked=False)
            for col in BLOCK["value_fields"].values():
                ws[f"{col}{row}"].protection = Protection(locked=False)
    buf = io.BytesIO()
    wb.save(buf)
    template_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")

    resp = client.post("/", json={
        "template": template_b64,
        "rooms": [
            {"floor": "1", "room_name": "和室", "measurements": {
                "floor_x": {"select": "←", "diff": "3", "distance": "1500"},
            }},
            {"floor": "2", "room_name": "洋室1", "measurements": {}},
        ],
    })
    assert resp.status_code == 200
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]
    # 1 部屋目（和室）が先頭ブロックに保存されていること。
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value == "和室"
    assert ws[_select_cell(0, "floor_x")].value == "←"
    assert ws[_value_cell(0, "floor_x", "diff")].value == 3
    # 2 部屋目（洋室1）は次のブロックへ。
    assert ws[f"{BLOCK['room_name_col']}{STARTS[1]}"].value == "洋室1"


def test_locked_cells_in_protected_sheet_preserve_existing_content(client):
    # シート保護が有効でロックされたセルはクリアフェーズでも既存値を保持する。
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.protection.sheet = True
    ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value = "（例）"  # 入力例の既存値
    ws[MAPPING["property_name_cell"]].protection = Protection(locked=False)
    buf = io.BytesIO()
    wb.save(buf)
    template_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")

    resp = client.post("/", json={
        "template": template_b64,
        "property_name": "サンプル物件",
        "rooms": [],
    })
    assert resp.status_code == 200
    ws = _decode_workbook(resp.get_json()["fileData"])[SHEET]
    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value == "（例）"
    assert ws[MAPPING["property_name_cell"]].value == "サンプル物件"


# --- Bad / missing input ----------------------------------------------------

def test_post_without_body_returns_400(client):
    resp = client.post("/", data="", content_type="application/json")

    assert resp.status_code == 400
    assert resp.headers["Access-Control-Allow-Origin"] == "*"
    assert resp.get_json() == {"error": "No data provided"}


def test_post_empty_json_object_returns_400(client):
    # An empty object is falsy, so the function treats it as "no data".
    resp = client.post("/", json={})

    assert resp.status_code == 400
    assert resp.get_json() == {"error": "No data provided"}


def test_non_json_body_returns_400(client):
    # get_json(silent=True) yields None for unparseable bodies.
    resp = client.post("/", data="not-json", content_type="text/plain")

    assert resp.status_code == 400
    assert resp.get_json() == {"error": "No data provided"}


def test_post_without_template_returns_400(client):
    # Data is present but no template was supplied by the caller.
    resp = client.post("/", json={"property_name": "X", "rooms": []})

    assert resp.status_code == 400
    assert resp.headers["Access-Control-Allow-Origin"] == "*"
    assert resp.get_json() == {"error": "No template provided"}


# --- Error handling ---------------------------------------------------------

def test_invalid_template_bytes_returns_500_with_message(client):
    # A non-xlsx Base64 blob makes load_workbook raise, exercising the
    # ``except`` branch that returns a 500 with the error text.
    bogus = base64.b64encode(b"not a real workbook").decode("utf-8")

    resp = client.post("/", json={"rooms": [], "template": bogus})

    assert resp.status_code == 500
    assert resp.headers["Access-Control-Allow-Origin"] == "*"
    assert resp.get_json()["error"]


def test_template_without_expected_sheet_returns_500(client):
    # A valid workbook that lacks the 傾斜測定 sheet raises a KeyError.
    resp = client.post("/", json={
        "rooms": [{"floor": "1", "room_name": "LDK", "measurements": {}}],
        "template": _make_template_b64(sheet_name="別のシート"),
    })

    assert resp.status_code == 500
    assert SHEET in resp.get_json()["error"]
