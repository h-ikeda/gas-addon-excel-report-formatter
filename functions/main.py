import functions_framework
import base64
import io
import json
import os
import unicodedata
from openpyxl import load_workbook
from flask import jsonify

# 書き込み位置（セルマッピング）は mapping.json に切り出している。
# フォーマットが微修正された場合は、原則 mapping.json を編集するだけで追従できる。
_MAPPING_PATH = os.path.join(os.path.dirname(__file__), 'mapping.json')
_MAPPING = None


def _load_mapping():
    # mapping.json は静的な設定ファイルなので、一度読み込んだらキャッシュする。
    # Cloud Functions のインスタンスはリクエスト間で再利用されるため、毎回の
    # ディスク I/O を避けられる。
    global _MAPPING
    if _MAPPING is None:
        with open(_MAPPING_PATH, encoding='utf-8') as f:
            _MAPPING = json.load(f)
    return _MAPPING


def _to_cell_value(value):
    """フォームから来る値を Excel セル向けに整える。

    数値として解釈できる文字列は数値に変換する（換算計測値 AJ の数式
    =1000*AC/AG が正しく計算されるよう、測定値の差・距離などは数値で書き込む）。
    空文字・None は None（空セル）として扱う。
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text == '':
        return None
    # モバイルの日本語入力では全角の数字・記号（－ ． など）が混じりやすい。
    # NFKC 正規化で半角へ寄せてから数値判定することで、Excel 側で数値として
    # 認識されず数式が計算されない事故を防ぐ。
    text = unicodedata.normalize('NFKC', text)
    try:
        if text.lstrip('-').isdigit():
            return int(text)
        return float(text)
    except ValueError:
        return text


def _clear_data_cells(ws, mapping):
    """全ブロックの「記入欄」を空にする。

    雛形には記入例（(例)ＬＤＫ など）が入っていることがあり、利用者が記入
    しなかった欄に例の値が残ると誤った報告書になる。そこで mapping.json が
    記入欄として定義しているセル（物件名・階数・部屋名・各計測点の選択欄と
    数値欄）だけを一度クリアしてから書き込む。印字済みのラベル・区切りの
    「/」・分母 1000・換算計測値の数式などは mapping に含まれないため消さない。
    """
    block = mapping['room_block']
    value_fields = block['value_fields']

    ws[mapping['property_name_cell']] = None
    for start_row in mapping['block_start_rows']:
        ws[f"{block['floor_col']}{start_row}"] = None
        ws[f"{block['room_name_col']}{start_row}"] = None
        for m in block['measurements']:
            row = start_row + m['row_offset']
            select = m.get('select')
            if select:
                ws[f"{select['col']}{row}"] = None
            for col in value_fields.values():
                ws[f'{col}{row}'] = None


def _write_room(ws, mapping, start_row, room):
    """1 部屋ぶんのデータを、先頭行 start_row のブロックへ書き込む。"""
    # 不正な形（None や辞書以外）の要素は無視し、AttributeError を防ぐ。
    if not isinstance(room, dict):
        return

    block = mapping['room_block']

    ws[f"{block['floor_col']}{start_row}"] = _to_cell_value(room.get('floor'))
    ws[f"{block['room_name_col']}{start_row}"] = _to_cell_value(room.get('room_name'))

    value_fields = block['value_fields']
    measurements = room.get('measurements') or {}
    for m in block['measurements']:
        data = measurements.get(m['key'])
        if not isinstance(data, dict):
            continue
        row = start_row + m['row_offset']
        # 選択欄（傾斜方向／測定した壁・柱）。雛形のプルダウン候補をそのまま
        # 書き込むため、数値化や正規化はせず文字列のまま入れる。
        select = m.get('select')
        if select and data.get('select') not in (None, ''):
            ws[f"{select['col']}{row}"] = data['select']
        # 数値欄（水平器計測値・測定値の差・距離）。
        for field_key, col in value_fields.items():
            if field_key in data:
                ws[f'{col}{row}'] = _to_cell_value(data.get(field_key))


@functions_framework.http
def generate_excel(request):
    if request.method == 'OPTIONS':
        headers = {
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'POST',
            'Access-Control-Allow-Headers': 'Content-Type',
            'Access-Control-Max-Age': '3600'
        }
        return ('', 204, headers)

    headers = {'Access-Control-Allow-Origin': '*'}

    try:
        request_json = request.get_json(silent=True)
        if not request_json:
            return jsonify({"error": "No data provided"}), 400, headers

        # 雛形（テンプレート）は GAS 側が Google Drive から読み取り、Base64 で
        # 渡してくる。社外秘フォーマットを Cloud Function に同梱しないことで、
        # フォーマット更新時の再デプロイを不要にし、ファイルのアクセス範囲を
        # 社内の閲覧可能者だけに閉じる。
        template_b64 = request_json.get('template')
        if not template_b64:
            return jsonify({"error": "No template provided"}), 400, headers

        mapping = _load_mapping()
        rooms = request_json.get('rooms') or []
        block_start_rows = mapping['block_start_rows']
        if len(rooms) > len(block_start_rows):
            return jsonify({
                "error": (
                    f"部屋数が雛形の上限（{len(block_start_rows)} 部屋）を超えています。"
                    "雛形にブロックを追加し、mapping.json の block_start_rows を更新してください。"
                )
            }), 400, headers

        wb = load_workbook(filename=io.BytesIO(base64.b64decode(template_b64)))
        ws = wb[mapping['sheet_name']]

        # 雛形の記入例・前回値が残らないよう、記入欄を一度クリアする。
        _clear_data_cells(ws, mapping)

        # 物件名（共通項目）
        if 'property_name' in request_json:
            ws[mapping['property_name_cell']] = _to_cell_value(
                request_json.get('property_name'))

        # 各部屋のデータを、対応するブロックへ順に書き込む。
        for i, room in enumerate(rooms):
            _write_room(ws, mapping, block_start_rows[i], room)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_excel = base64.b64encode(output.read()).decode('utf-8')

        return jsonify({
            "status": "success",
            "fileName": "傾斜測定報告書.xlsx",
            "fileData": encoded_excel
        }), 200, headers

    except Exception as e:
        return jsonify({"error": str(e)}), 500, headers
